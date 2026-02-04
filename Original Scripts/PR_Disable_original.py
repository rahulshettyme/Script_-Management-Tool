# EXPECTED_INPUT_COLUMNS: id, name, deletion response, deletion status, request id

# AI Updated Script - 2026-01-31 14:51:35 IST
# ======================================================
# 📘 RS_stop_pr.py
# Author: Rahul Shetty (Standard Format)
# Purpose:
#   - Standalone script only
#   - Read Environment_Details & Plot_details
#   - Generate token
#   - PHASE 1: Send delete requests for all rows
#   - PHASE 2: Status check for all collected requestIds
#   - Update Excel (replace sheet)
#   - Execution time tracking
# ======================================================
# 🔧 Imports
import os
import time
import argparse
from datetime import datetime
import requests
import pandas as pd
from openpyxl import load_workbook
from RS_access_token_generate import get_bearer_token

# ==============================
# 📄 File paths & settings (STANDARD FORMAT)
# ==============================
file_path = r"C:\Users\cropin\Documents\Important\Excel file\Delete PR.xlsx"
sheet_name = "Plot_details"
env_sheet_name = "Environment_Details"

print(f"📂 Loading Excel: {file_path}")

# ==============================
# 🔐 Get Access Token
# ==============================
print("🔄 Requesting access token...")
token = get_bearer_token(file_path)
if not token:
    print("❌ Failed to retrieve token. Exiting.")
    raise SystemExit(1)

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}

# ==============================
# 🌍 Read environment base URL (openpyxl)
# ==============================
wb = load_workbook(file_path, data_only=True)
if env_sheet_name not in wb.sheetnames:
    raise RuntimeError(f"❌ Sheet '{env_sheet_name}' not found in workbook")

env_sheet = wb[env_sheet_name]
env_key = None
for r in range(2, env_sheet.max_row + 1):
    raw = env_sheet.cell(row=r, column=1).value
    if raw and str(raw).strip().lower() == "environment":
        env_key = str(env_sheet.cell(row=r, column=2).value).strip()
        break

env_url = None
if env_key:
    for r in range(2, env_sheet.max_row + 1):
        raw = env_sheet.cell(row=r, column=1).value
        if raw and str(raw).strip().lower() == env_key.lower():
            env_url = str(env_sheet.cell(row=r, column=2).value).strip()
            break

if not env_url:
    raise RuntimeError("❌ Base URL for environment not found in Environment_Details sheet")

env_url = env_url.rstrip('/')
print(f"🌍 Using Base URL: {env_url}")

# Step 1 API Path Definition: POST /services/farm/api/intelligence/croppable-areas/request
DELETE_PLOT_API = f"{env_url}/services/farm/api/intelligence/croppable-areas/request"

# Step 2 API Path Definition: GET /services/farm/api/intelligence/croppable-areas/request/status
STATUS_CHECK_API = f"{env_url}/services/farm/api/intelligence/croppable-areas/request/status?requestId={{}}"

# ==============================
# 🚀 Load Plot_details sheet into DataFrame
# ==============================
df = pd.read_excel(file_path, sheet_name=sheet_name, engine="openpyxl")

# Normalize column names: strip spaces and lowercase
df.columns = [c.strip().lower() for c in df.columns]
print("Columns in Excel:", df.columns.tolist())
print("First few rows:\n", df.head())

# Ensure status/response/request id columns exist and are string dtype
# Required Columns: id, name, deletion response, deletion status, request id
for col in ["deletion response", "deletion status", "request id"]:
    if col not in df.columns:
        df[col] = ""
    else:
        df[col] = df[col].astype(str)

# Ensure 'id' column exists (required)
if "id" not in df.columns:
    raise RuntimeError("❌ Required column 'id' not found in Plot_details sheet")

# ==============================
# 💾 Save DataFrame back to Excel (with retries & backup)
# ==============================
def save_df_to_excel(df_to_save, file_path, sheet_name=sheet_name, max_retries=3):
    # Step 3 [LOGIC]: Saves the updated DataFrame back to the 'Plot_details' sheet, replacing the sheet.
    attempt = 0
    while attempt < max_retries:
        try:
            with pd.ExcelWriter(file_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                df_to_save.to_excel(writer, sheet_name, index=False)
            print(f"✅ Excel updated: {file_path} (sheet: {sheet_name})")
            return True
        except PermissionError:
            attempt += 1
            print(f"⚠️ Permission denied when saving Excel. Ensure the file is closed. Retry {attempt}/{max_retries} ...")
            time.sleep(4)
        except FileNotFoundError:
            try:
                with pd.ExcelWriter(file_path, engine="openpyxl", mode="w") as writer:
                    df_to_save.to_excel(writer, sheet_name, index=False)
                print(f"✅ Excel created and saved: {file_path} (sheet: {sheet_name})")
                return True
            except Exception as e:
                print(f"❌ Failed to create Excel: {e}")
                break
        except Exception as e:
            print(f"❌ Unexpected error while saving Excel: {e}")
            break

    # fallback backup
    backup_path = file_path.replace(".xlsx", f"_backup_{int(time.time())}.xlsx")
    try:
        with pd.ExcelWriter(backup_path, engine="openpyxl", mode="w") as writer:
            df_to_save.to_excel(writer, sheet_name, index=False)
        print(f"❌ Could not save original file. Saved backup: {backup_path}")
        return False
    except Exception as e:
        print(f"❌ Failed to save backup file as well: {e}")
        return False

# ==============================
# 🔁 PHASE 1: Send deletes for all rows (collect request ids)
# Step 1: Send Plot Deletion Request
# ==============================
def phase1_send_deletes(df_in, headers, delete_api=DELETE_PLOT_API, per_call_sleep=0.4):
    print("===========================================")
    print("🔁 PHASE 1: Sending DELETE request for all rows")
    print("===========================================")

    for idx, row in df_in.iterrows():
        plot_id = row.get("id", "")
        if pd.isna(plot_id) or str(plot_id).strip() == "":
            print(f"⚠️ Row {idx+1}: Empty ID → skipping")
            df_in.at[idx, "deletion response"] = "Skipped: empty id"
            df_in.at[idx, "deletion status"] = "Skipped"
            df_in.at[idx, "request id"] = ""
            continue

        print(f"🧭 Row {idx+1}: Sending delete for Plot ID {plot_id}")
        try:
            # API Call: POST /services/farm/api/intelligence/croppable-areas/request, Payload: [df_in['id']]
            resp = requests.post(delete_api, json=[plot_id], headers=headers, timeout=60)
        except Exception as e:
            df_in.at[idx, "deletion response"] = f"Exception: {e}"
            df_in.at[idx, "deletion status"] = "Delete Failed"
            df_in.at[idx, "request id"] = ""
            print(f"    ❌ Exception during delete call: {e}")
            time.sleep(per_call_sleep)
            continue

        if resp.status_code == 200:
            try:
                resp_json = resp.json()
            except Exception:
                resp_json = resp.text

            # Expected Response: df_in['deletion response']
            df_in.at[idx, "deletion response"] = str(resp_json)

            # attempt to extract request id (Preserve complex extraction logic)
            req_id = ""
            if isinstance(resp_json, dict):
                req_id = resp_json.get("id") or resp_json.get("requestId") or resp_json.get("request_id") or ""
                if not req_id:
                    # inspect nested dicts/lists
                    for v in resp_json.values():
                        if isinstance(v, dict):
                            req_id = v.get("id") or v.get("requestId") or ""
                            if req_id:
                                break
            elif isinstance(resp_json, list) and len(resp_json) > 0 and isinstance(resp_json[0], dict):
                req_id = resp_json[0].get("id") or resp_json[0].get("requestId") or ""

            # Expected Response: df_in['request id']
            df_in.at[idx, "request id"] = req_id or ""
            print(f"    ✔️ Delete queued. Request Id: {req_id or 'N/A'}")
        else:
            df_in.at[idx, "deletion response"] = f"Error {resp.status_code}: {resp.text}"
            df_in.at[idx, "deletion status"] = "Delete Failed"
            df_in.at[idx, "request id"] = ""
            print(f"    ❌ Delete failed (HTTP {resp.status_code}) for Plot ID {plot_id}")

        time.sleep(per_call_sleep)

    return df_in

# ==============================
# 🔁 PHASE 2: Check status for all collected request ids
# Step 2: Check Plot Deletion Status
# ==============================
def phase2_check_status(df_in, headers, status_api_template=STATUS_CHECK_API, post_delete_pause=8, per_status_sleep=0.4, max_status_attempts=1):
    print("\n⏳ Waiting fixed period before status checks...")
    time.sleep(post_delete_pause)

    print("===========================================")
    print("🔁 PHASE 2: Checking STATUS for all rows")
    print("===========================================")

    for idx, row in df_in.iterrows():
        plot_id = row.get("id", "")
        req_id = row.get("request id", "")

        if pd.isna(plot_id) or str(plot_id).strip() == "":
            continue  # skipped in phase1

        if not req_id or str(req_id).strip() == "":
            # Fallback handling: Payload Example: df_in['id'] for fallback
            current_resp = str(row.get("deletion response", ""))
            if "Error" in current_resp or "Exception" in current_resp:
                df_in.at[idx, "deletion status"] = "Delete failed - no request id"
                print(f"⚠️ Row {idx+1}: No request id; delete failed earlier.")
            else:
                try:
                    print(f"🔎 Row {idx+1}: No request id; attempting fallback status check using Plot ID {plot_id}")
                    # API Call: GET /services/farm/api/intelligence/croppable-areas/request/status?requestId={plot_id}
                    fallback_resp = requests.get(status_api_template.format(plot_id), headers=headers, timeout=40)
                    if fallback_resp.status_code == 200:
                        try:
                            fallback_json = fallback_resp.json()
                        except Exception:
                            fallback_json = fallback_resp.text
                        # Expected Response: df_in['deletion status']
                        df_in.at[idx, "deletion status"] = str(fallback_json)
                        print(f"    🔄 Fallback status returned")
                    else:
                        df_in.at[idx, "deletion status"] = f"No request id; fallback error {fallback_resp.status_code}"
                        print(f"    ❌ Fallback status failed: {fallback_resp.status_code}")
                except Exception as e:
                    df_in.at[idx, "deletion status"] = f"No request id; fallback exception: {e}"
                    print(f"    ❌ Exception during fallback: {e}")
            time.sleep(per_status_sleep)
            continue

        # Primary handling: Payload Example: requestId = df_in['request id']
        status_value = None
        for attempt in range(1, max_status_attempts + 1):
            try:
                # API Call: GET /services/farm/api/intelligence/croppable-areas/request/status?requestId={req_id}
                status_resp = requests.get(status_api_template.format(req_id), headers=headers, timeout=60)
            except Exception as e:
                status_value = f"Exception: {e}"
                print(f"    ❌ Exception while checking status for RequestId {req_id}: {e}")
                break

            if status_resp.status_code == 200:
                try:
                    status_json = status_resp.json()
                except Exception:
                    status_json = status_resp.text
                status_value = str(status_json)
                print(f"    🔄 Row {idx+1}: Status retrieved")
                break
            else:
                status_value = f"Error {status_resp.status_code}: {status_resp.text}"
                print(f"    ❌ Status check attempt {attempt} failed for RequestId {req_id} (HTTP {status_resp.status_code})")
                if attempt < max_status_attempts:
                    time.sleep(per_status_sleep)

        # Expected Response: df_in['deletion status']
        df_in.at[idx, "deletion status"] = status_value or "No status returned"
        time.sleep(per_status_sleep)

    return df_in

# ==============================
# 🚀 MAIN (Standalone)
# ==============================
def main():
    start_time = datetime.now()
    print("🔄 Starting Plot deletion process...")

    # Phase 1: send deletes
    updated_df = phase1_send_deletes(df, headers, delete_api=DELETE_PLOT_API, per_call_sleep=0.4)

    # Phase 2: status checks
    updated_df = phase2_check_status(updated_df, headers, status_api_template=STATUS_CHECK_API, post_delete_pause=8, per_status_sleep=0.4, max_status_attempts=1)

    # Save results back to Excel (Step 3)
    # Ensure we write columns names lowercased (consistent with standard)
    updated_df.columns = [c.strip().lower() for c in updated_df.columns]
    saved = save_df_to_excel(updated_df, file_path, sheet_name)
    if not saved:
        print("⚠️ Could not save to original file; backup created.")

    end_time = datetime.now()
    elapsed = end_time - start_time
    print("======================================================")
    print(f"Start Time : {start_time}")
    print(f"End Time   : {end_time}")
    print(f"Elapsed    : {elapsed}")
    print("======================================================")

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="Plot deletion processor (standalone)")
    parser.add_argument("--file", "-f", default=file_path, help="Path to Excel file")
    parser.add_argument("--sheet", "-s", default=sheet_name, help="Sheet name containing plots")
    args = parser.parse_args()

    # Allow overriding file_path/sheet via CLI while keeping same code flow
    file_override = args.file
    sheet_override = args.sheet
    if file_override and file_override != file_path:
        file_path = file_override
        print(f"📂 Using file override: {file_path}")
        # reload workbook & df based on override
        wb = load_workbook(file_path, data_only=True)
        if env_sheet_name not in wb.sheetnames:
            raise RuntimeError(f"❌ Sheet '{env_sheet_name}' not found in workbook")
        env_sheet = wb[env_sheet_name]
        env_key = None
        for r in range(2, env_sheet.max_row + 1):
            raw = env_sheet.cell(row=r, column=1).value
            if raw and str(raw).strip().lower() == "environment":
                env_key = str(env_sheet.cell(row=r, column=2).value).strip()
                break
        env_url = None
        if env_key:
            for r in range(2, env_sheet.max_row + 1):
                raw = env_sheet.cell(row=r, column=1).value
                if raw and str(raw).strip().lower() == env_key.lower():
                    env_url = str(env_sheet.cell(row=r, column=2).value).strip()
                    break
        if not env_url:
            raise RuntimeError("❌ Base URL for environment not found in Environment_Details sheet")
        env_url = env_url.rstrip('/')
        
        # Re-define APIs based on new environment URL
        DELETE_PLOT_API = f"{env_url}/services/farm/api/intelligence/croppable-areas/request"
        STATUS_CHECK_API = f"{env_url}/services/farm/api/intelligence/croppable-areas/request/status?requestId={{}}"

        df = pd.read_excel(file_path, sheet_name=sheet_override, engine="openpyxl")
        df.columns = [c.strip().lower() for c in df.columns]
        for col in ["deletion response", "deletion status", "request id"]:
            if col not in df.columns:
                df[col] = ""
            else:
                df[col] = df[col].astype(str)

    main()