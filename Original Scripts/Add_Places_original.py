# ======================================================
# 📘 RS_Add_Places.py
# Author: Rahul Shetty (updated)
# Purpose:
#  - Use RS_address_generate.get_location_details(...) exactly like RS_Add_Asset
#  - If latitude & longitude provided use them, else use address column
#  - Remove internal Google Maps logic and call address generator exactly like asset script
#  - Ensure both 'status' and 'response' reliably written to Excel
# ======================================================

# ==============================
# 🔌 Imports
# ==============================
import time
import requests
import json
import math
from datetime import datetime
import pandas as pd
from openpyxl import load_workbook
import openpyxl

from RS_access_token_generate import get_bearer_token
from RS_address_generate import get_location_details

# ==============================
# 🕒 Start time
# ==============================
start_time = datetime.now()

# ==============================
# 📄 File paths & settings
# ==============================
file_path = r"C:\Users\cropin\Documents\Important\Excel file\Add Place.xlsx"
sheet_name = "Place_details"
env_sheet_name = "Environment_Details"

print(f"📂 Loading Excel: {file_path}")

# ==============================
# 🔐 Get Access Token
# ==============================
print("🔄 Requesting access token...")
token = get_bearer_token(file_path)
if not token:
    print("❌ Failed to retrieve token. Exiting.")
    raise SystemExit(1)

headers = {
    "Authorization": f"Bearer {token}",
    "Content-Type": "application/json"
}
print("✅ Token retrieved successfully!")

# ==============================
# 🌍 Read environment base URL
# ==============================
# Use openpyxl to read Environment_Details (keeps compatibility with your workbook)
wb = load_workbook(file_path, data_only=True)
if env_sheet_name not in wb.sheetnames:
    raise RuntimeError(f"❌ Sheet '{env_sheet_name}' not found in workbook")

env_sheet = wb[env_sheet_name]
env_key = None
for r in range(2, env_sheet.max_row + 1):
    raw = env_sheet.cell(row=r, column=1).value
    if raw and str(raw).strip().lower() == "environment":
        env_key = str(env_sheet.cell(row=r, column=2).value).strip()
        break

env_url = None
if env_key:
    for r in range(2, env_sheet.max_row + 1):
        raw = env_sheet.cell(row=r, column=1).value
        if raw and str(raw).strip().lower() == env_key.lower():
            env_url = str(env_sheet.cell(row=r, column=2).value).strip()
            break

if not env_url:
    raise RuntimeError("❌ Base URL for environment not found in Environment_Details sheet")

print(f"🌍 Using Base URL: {env_url.rstrip('/')}")
place_url = f"{env_url.rstrip('/')}/services/farm/api/place"

# ==============================
# 🚀 Load Place Details sheet into DataFrame
# ==============================
df = pd.read_excel(file_path, sheet_name=sheet_name)

# Normalize column names: strip spaces and lowercase
df.columns = [c.strip().lower() for c in df.columns]
print("Columns in Excel:", df.columns.tolist())
print("First few rows:\n", df.head())

# Ensure status and response columns exist and are string dtype
for col in ["status", "response"]:
    if col not in df.columns:
        df[col] = ""
    else:
        df[col] = df[col].astype(str)

# ==============================
# 🔧 Helper: validate coordinates
# ==============================
def is_valid_coord(v):
    """Return True if v is a finite number (not None / empty / NaN)."""
    if v is None:
        return False
    if isinstance(v, str) and v.strip() == "":
        return False
    try:
        num = float(v)
        return math.isfinite(num)
    except Exception:
        return False

# ==============================
# 🏁 Main loop: process each row
# ==============================
for index, row in df.iterrows():
    try:
        # 🔹 Read input fields from Excel (normalized)
        name = str(row.get("name", "")).strip()
        place_type = str(row.get("type", "")).strip().upper()
        lat_raw = row.get("latitude", None)
        lng_raw = row.get("longitude", None)
        address_text = ""
        if row.get("address") is not None:
            address_text = str(row.get("address")).strip()

        # Basic validation: name and type required
        if not name or not place_type:
            df.at[index, "status"] = "⚠️ Skipped - Missing required fields (name/type)"
            df.at[index, "response"] = ""
            print(f"⚠️ Skipped row {index+2} due to missing name/type")
            continue

        use_coords = is_valid_coord(lat_raw) and is_valid_coord(lng_raw)
        lat_val = float(lat_raw) if is_valid_coord(lat_raw) else None
        lng_val = float(lng_raw) if is_valid_coord(lng_raw) else None

        print(f"🔄 Processing: {name} | Type: {place_type} | use_coords: {use_coords} | lat: {lat_val} lng: {lng_val}")

        # ==============================
        # 🔎 Call address generator (exactly like asset script)
        # ==============================
        addr = None
        try:
            # Primary call: (address, lat, lon) — same as asset script
            print(f"➡️ Calling get_location_details(address, lat, lon) with -> address: '{address_text[:80]}' lat: {lat_val} lon: {lng_val}")
            addr = get_location_details(address_text, lat_val, lng_val)
            print("🔍 Returned (primary):", addr)
        except TypeError as e:
            print("⚠️ Primary call raised TypeError - trying alternative signatures. Error:", e)
            try:
                # try coords-only if coords present
                if use_coords:
                    print(f"➡️ Trying get_location_details(lat, lon) -> {lat_val}, {lng_val}")
                    addr = get_location_details(lat_val, lng_val)
                    print("🔍 Returned (alt coords-only):", addr)
                # try named arguments
                if not addr:
                    print(f"➡️ Trying get_location_details(address=..., latitude=..., longitude=...)")
                    addr = get_location_details(address=address_text, latitude=lat_val, longitude=lng_val)
                    print("🔍 Returned (alt named):", addr)
            except Exception as e2:
                print("❌ All fallback address calls failed:", e2)
                addr = None
        except Exception as ex:
            print("❌ Unexpected error while calling get_location_details:", ex)
            addr = None

        # Final fallback: only address_text
        if not addr and address_text:
            try:
                print("➡️ Final fallback: calling get_location_details(address_text) only")
                addr = get_location_details(address_text)
                print("🔍 Returned (final fallback):", addr)
            except Exception as e:
                print("❌ Final fallback also failed:", e)
                addr = None

        if not addr:
            df.at[index, "status"] = "❌ Failed: No address"
            df.at[index, "response"] = "Address generator returned empty/falsy payload"
            print(f"❌ Address not found for {name} (see debug above)")
            # continue to next row
            continue

        # Ensure returned address contains lat/lon if possible
        lat_from_addr = None
        lon_from_addr = None
        try:
            if isinstance(addr, dict):
                lat_from_addr = float(addr.get("latitude")) if addr.get("latitude") is not None else None
                lon_from_addr = float(addr.get("longitude")) if addr.get("longitude") is not None else None
        except Exception:
            lat_from_addr = None
            lon_from_addr = None

        # ==============================
        # 🧩 Prepare payload for Add Place API
        # ==============================
        payload = {
            "name": name,
            "type": place_type,
            "subType": None,
            "capacity": {"count": 10},
            "address": addr,
            "areaAudit": None,
            "auditedArea": None,
            "visibility": True,
            "latitude": lat_from_addr if lat_from_addr is not None else lat_val,
            "longitude": lon_from_addr if lon_from_addr is not None else lng_val,
            "data": None,
            "images": None
        }

        # ==============================
        # 🚀 Call Add Place API
        # ==============================
        print(f"➡️ Calling Add Place API for {name} ...")
        try:
            response = requests.post(place_url, headers=headers, json=payload)
            response_text = str(response.text)
            print("🔁 API code:", response.status_code)
        except Exception as api_ex:
            response = None
            response_text = f"API call failed: {str(api_ex)}"
            print("❌ API call exception:", api_ex)

        # Ensure we always store string for response to avoid Excel <NA> problems
        df.at[index, "response"] = response_text
        if response is not None and response.status_code in (200, 201):
            df.at[index, "status"] = "✅ Success"
        elif response is not None:
            df.at[index, "status"] = f"❌ Failed: {response.status_code}"
        else:
            df.at[index, "status"] = "❌ Failed: API error"

        print(f"✅ Completed row: {name} | status: {df.at[index,'status']}")

    except Exception as e:
        df.at[index, "status"] = f"⚠️ Error: {str(e)}"
        df.at[index, "response"] = ""
        print(f"⚠️ Error for row {index + 2}: {str(e)}")

    # polite pause to avoid hitting rate limits
    time.sleep(0.3)

# ==============================
# 💾 Save Results Back to Excel (robust - remove & append)
# ==============================
# Convert response/status to string to avoid mixed-type save issues
if "response" in df.columns:
    df["response"] = df["response"].astype(str)
if "status" in df.columns:
    df["status"] = df["status"].astype(str)

# Load workbook, remove existing sheet (if exists), save, then append new sheet
book = load_workbook(file_path)
if sheet_name in book.sheetnames:
    del book[sheet_name]
    book.save(file_path)  # save after deletion so writer can append fresh

with pd.ExcelWriter(file_path, engine="openpyxl", mode="a") as writer:
    df.to_excel(writer, sheet_name=sheet_name, index=False)

print("\n📥 Excel updated & saved.")

# Optional quick verification: read back first 3 rows from sheet and print status/response cols
verify_df = pd.read_excel(file_path, sheet_name=sheet_name, usecols=["name", "status", "response"])
print("\n🔎 Verification (first 3 rows):")
print(verify_df.head(3).to_string(index=False))

# ==============================
# 🏁 End time / Execution summary
# ==============================
end_time = datetime.now()
print("\n🏁 Execution completed.")
print(f"Start Time : {start_time}")
print(f"End Time   : {end_time}")
print(f"Elapsed    : {end_time - start_time}")
